r"""
Generate `docs/concepts/persona-capabilities/test-scenarios.xlsx` —
the manual-comparison checklist for the Phase A-K scenario suite.

Layout: single sheet, sections stacked vertically. Each section is one
scenario. Per scenario:

  Title row (merged across both columns, bold + filled)
  Header row ("Flow step" | "Status")
  Data rows
  Blank separator row

Run from the repo root:
  python tools/test-mcp/generate_test_scenarios_xlsx.py

Idempotent — overwrites the destination file.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
DEST = REPO_ROOT / "docs" / "concepts" / "persona-capabilities" / "test-scenarios.xlsx"


# ---------------------------------------------------------------------------
# Scenario data
# ---------------------------------------------------------------------------

# Common flow steps shared across most build-then-promote scenarios.
# Specific scenarios extend this with their own gates.
def common_steps() -> list[tuple[str, str]]:
    return [
        ("Preflight — /health probe returns ok", "ok"),
        ("startBuildFromIntent → session created, persona row enabled=false", "ok"),
        ("LLM emits clarifying questions, driver answers via cellKey-keyed batch", "ok"),
        ("All gates open, LLM emits agent_ir → phase: draft_ready", "ok"),
        ("wait_for_agent_ir defensive poll succeeds", "ok"),
        ("triggerBuildTest → tool tests run, phase: test_complete", "ok"),
        ("promoteBuildDraft → single-tx commit; persona enabled=true", "ok"),
    ]


# NOTE: each scenario carries the EXACT prompt string the driver passes
# to `startBuildFromIntent`. Sourced verbatim from the matching
# `tools/test-mcp/e2e_phase_*.py` driver — keep these in sync when a
# driver's INTENT changes (regenerate the xlsx after any driver edit).

SCENARIOS: list[dict] = [
    {
        "title": "Phase A.1 — Inbox Triage",
        "summary": "2 use_cases both event-triggered on `gmail.message.received` (parallel, not chained).",
        "introduced": "C6 (2026-04-27)",
        "driver": "tools/test-mcp/e2e_phase_a.py --scenario inbox",
        "intent": (
            "Watch my Gmail inbox. On every new message, classify it as urgent / "
            "followup / fyi. When the classification is 'urgent', also draft a "
            "short reply for me to review before sending. Both capabilities "
            "trigger directly on a new Gmail message arriving — the draft "
            "capability does NOT chain off the classifier; it runs in parallel "
            "and only emits its draft when its own classifier-style check judges "
            "the message urgent."
        ),
        "steps": common_steps() + [
            ("Acceptance: 2 use_cases produced", "ok"),
            ("Acceptance: 2 event_listener triggers on `gmail.message.received`", "ok"),
            ("Acceptance: 1 assertion / 3 subs / 3 tools / 2 triggers landed", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase A.2 — Project Coordinator",
        "summary": "4 use_cases — 3 collectors (Linear/GitHub/Calendar) + 1 digest assembler chained on `*.brief.ready` events. Validates rule D (producer/publisher chain split).",
        "introduced": "C6 (2026-04-27)",
        "driver": "tools/test-mcp/e2e_phase_a.py --scenario coordinator",
        "intent": (
            "Every Monday at 8am local time, build a single weekly digest "
            "covering three sources: my Linear assigned issues, my GitHub pull "
            "requests awaiting review, and today's Google Calendar events. Each "
            "source is its own capability so I can turn off any one "
            "independently. Save the digest as a single markdown file in my "
            "local drive — never email or message."
        ),
        "steps": common_steps() + [
            ("Acceptance: 4 use_cases produced (rule D fired — collectors split + assembler)", "ok"),
            ("Acceptance: 4 triggers landed", "ok"),
            ("Acceptance: brief.ready event chain wired between collectors and assembler", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase B — Chained personas X→Y→Z",
        "summary": "3 personas: HN scraper → Summarizer → Archivist. Validates cross-persona event chain (auto-`source_filter='*'` for inbound listens).",
        "introduced": "C6 (2026-04-27)",
        "driver": "tools/test-mcp/e2e_phase_b.py",
        "intent": (
            "[Persona X — head]\n"
            "Be a Hacker News digest scraper. Once an hour, fetch the top 3 "
            "stories from the Hacker News RSS feed (https://news.ycombinator.com/rss) "
            "and emit each as a draft news item for downstream personas to "
            "summarize. Single capability, schedule-triggered. (+event-name "
            "discipline lock for `news.draft.captured`.)\n"
            "\n"
            "[Persona Y — middle]\n"
            "Listen for `news.draft.captured` events. For each draft, write a "
            "2-sentence summary and emit it for the notes poster persona to "
            "archive. Emit `news.summary.ready` when the summary is ready. "
            "(+event-name discipline lock for both names.)\n"
            "\n"
            "[Persona Z — tail]\n"
            "Listen for `news.summary.ready` events. For each summary, append a "
            "markdown line to my local drive notes file at `news/digest.md`. "
            "(+event-name discipline lock for `news.summary.ready`.)"
        ),
        "steps": [
            ("Preflight — /health probe returns ok", "ok"),
            ("Build persona X (scraper) — schedule trigger, emits `news.draft.captured`", "ok"),
            ("Build persona Y (summarizer) — listens `news.draft.captured`, emits `news.summary.ready`", "ok"),
            ("Build persona Z (archivist) — listens `news.summary.ready`, writes to local-drive", "ok"),
            ("Promote each — auto-`source_filter='*'` defaulted for cross-persona listens", "ok"),
            ("Fire X manually → X executes", "ok"),
            ("Y picks up automatically (cross-persona event delivered)", "ok"),
            ("Z picks up automatically", "ok"),
            ("Acceptance: cascade.full_chain reports ok", "ok"),
            ("Cleanup: deleteAgent for X, Y, Z", "ok"),
        ],
    },
    {
        "title": "Phase C — Output diversity (vector_db scenario)",
        "summary": "Persona writes facts to `personas_vector_db` connector via `vector_db_insert` tool.",
        "introduced": "C6 (2026-04-27)",
        "driver": "tools/test-mcp/e2e_phase_c.py --scenario vector_db",
        "intent": (
            "Once a day at 8am, write a one-paragraph summary of yesterday's "
            "activity into my built-in vector knowledge base so I can search "
            "across past days later. Single capability, schedule-triggered. "
            "Output target: `personas_vector_db` connector — index the summary "
            "as a single document with date as the title."
        ),
        "steps": common_steps() + [
            ("Acceptance: persona uses `personas_vector_db` connector", "ok"),
            ("Acceptance: tool list includes `vector_db_insert` + `date_utilities`", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase C — Output diversity (notion scenario)",
        "summary": "Persona logs daily status to Notion page; uses `notion` + `gmail` connectors.",
        "introduced": "C6 (2026-04-27)",
        "driver": "tools/test-mcp/e2e_phase_c.py --scenario notion",
        "intent": (
            "Every morning at 9am, append a status page to my Notion workspace "
            "summarising overnight activity. Single capability, schedule-"
            "triggered. Output target: Notion — create one new page per run in "
            "a configured database."
        ),
        "steps": common_steps() + [
            ("Acceptance: persona uses `notion` + `gmail` connectors", "ok"),
            ("Acceptance: `notion_create_page` tool listed (+ web/gmail search)", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase C — Output diversity (github scenario)",
        "summary": "Weekly priority tracker creates GitHub issues; uses `github` + `personas_database`.",
        "introduced": "C6 (2026-04-27)",
        "driver": "tools/test-mcp/e2e_phase_c.py --scenario github",
        "intent": (
            "Every Monday at 7am, file a single GitHub issue summarising weekly "
            "priorities into a configured repo. Single capability, schedule-"
            "triggered. Output target: GitHub — issue creation."
        ),
        "steps": common_steps() + [
            ("Acceptance: persona uses `github` + `personas_database` connectors", "ok"),
            ("Acceptance: `github_create_issue` tool listed (+ list/search/sql)", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase C — Output diversity (titlebar scenario)",
        "summary": "Heartbeat monitor surfaces output to the desktop titlebar; no external connectors. Validates fix H (`notification_channel_array` v3.1 shape).",
        "introduced": "C6 (2026-04-27)",
        "driver": "tools/test-mcp/e2e_phase_c.py --scenario titlebar",
        "intent": (
            "Every hour on the hour, surface a desktop / titlebar notification "
            "with a single-sentence status update — NO file writes, NO event "
            "emissions, NO external connector. Just a notification."
        ),
        "steps": common_steps() + [
            ("Acceptance: no external connectors", "ok"),
            ("Acceptance: notification channel `built-in/titlebar` lands on persona row", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase D — auto_triage build-shape",
        "summary": "Verifies build prompt rule 21 fires `mode: \"auto_triage\"` and promote preserves it on `design_context.useCases[]`. Build-shape only — runtime verified by Phase D2.",
        "introduced": "C7 (2026-04-28)",
        "driver": "tools/test-mcp/e2e_phase_d.py",
        "intent": (
            "Once a day at 9am, scan my local-drive 'inbox' folder for new "
            "support emails I dropped there overnight. For each email, decide "
            "if it deserves a reply this morning by checking it against three "
            "decision principles: (1) the sender is a paying customer, (2) the "
            "subject mentions a P1/P2 incident, (3) it cites a broken feature. "
            "AUTO-TRIAGE the verdict — surface compliant ones in a titlebar "
            "notification, silently archive the rest. Use review_policy.mode = "
            "'auto_triage'. No human gate."
        ),
        "steps": common_steps() + [
            ("Acceptance: at least one UC has `review_policy.mode == \"auto_triage\"`", "ok"),
            ("Acceptance: `last_design_result.persona.decision_principles[]` non-empty (informational)", "ok"),
            ("Acceptance: 0 pending manual_reviews for the fresh persona (smoke check)", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase D2 — auto_triage runtime E2E",
        "summary": "Promotes a Phase-D-shaped persona, calls `synthesizeManualReview` (test bridge), polls until evaluator finalises verdict, asserts `policy_events` carries the `review.auto_triage.{approved|rejected|fallback}` audit tag.",
        "introduced": "C8 (2026-04-28)",
        "driver": "tools/test-mcp/e2e_phase_d2.py",
        "intent": (
            "Once a day at 9am, scan my local-drive 'inbox' folder for new "
            "support emails I dropped there overnight. For each email, decide "
            "if it deserves a reply this morning by checking it against three "
            "decision principles: (1) the sender is a paying customer, (2) the "
            "subject mentions a P1/P2 incident, (3) it cites a broken feature. "
            "AUTO-TRIAGE the verdict — surface compliant ones in a titlebar "
            "notification, silently archive the rest. Use review_policy.mode = "
            "'auto_triage'. No human gate.\n"
            "\n"
            "[After promote, the driver synthesizes a `manual_review` row via "
            "the C8 `synthesize_manual_review` test bridge command — title \"P1 "
            "incident — vault sync stuck for paying customer\", description "
            "matching all three decision principles, severity \"high\".]"
        ),
        "steps": common_steps() + [
            ("Find auto_triage UC on promoted persona's design_context", "ok"),
            ("synthesizeManualReview bridge call returns {reviewId, executionId}", "ok"),
            ("Poll listManualReviews until status leaves Pending (timeout 120s)", "ok"),
            ("Assert review row landed Approved / Rejected / Resolved", "ok"),
            ("getPolicyEventsForExecution returns at least one auto_triage audit row", "ok"),
            ("Latest run: verdict APPROVED in 8s, audit tag review.auto_triage.approved", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase E — Preload-error reload (deferred-manual)",
        "summary": "WebView2 stale-chunk recovery — verifies `vite:preloadError` listener reloads cleanly after a tauri-cli rebuild. Listener has 6 unit tests; live verification is a manual checklist.",
        "introduced": "C6 (2026-04-27)",
        "driver": "(manual — see C7-handoff-2026-04-28.md §Phase E)",
        "intent": (
            "(no build INTENT — Phase E is a manual checklist for the WebView2 "
            "stale-chunk reload path. Repro: launch dev with test-automation, "
            "navigate to a non-default route to force a lazy chunk load, touch "
            "any Rust source file to trigger a tauri-cli rebuild + restart, "
            "click around in the WebView once the backend is back up. Watch for "
            "`vite:preloadError — reloading to pick up fresh chunks` console "
            "log + clean reload. Verify 30s throttle by deleting a chunk file "
            "mid-flight — the second preloadError within 30s should log the "
            "throttle suppression instead of reloading.)"
        ),
        "steps": [
            ("Launch dev with test-automation; navigate to a non-default route (lazy chunk)", "deferred-manual"),
            ("Touch any Rust source file to force tauri-cli rebuild + restart", "deferred-manual"),
            ("After backend returns, click around in the WebView", "deferred-manual"),
            ("Watch for `vite:preloadError — reloading to pick up fresh chunks` log + clean reload", "deferred-manual"),
            ("Verify 30s throttle by deleting a chunk file mid-flight", "deferred-manual"),
            ("Listener unit tests (preloadErrorRecovery.test.ts, 6 cases)", "ok"),
        ],
    },
    {
        "title": "Phase F — Multi-language build (Czech preset)",
        "summary": "Verifies localised fields (mission/principles/decision_principles/operating_instructions) come back in target language while UC ids and trigger_type strings stay English.",
        "introduced": "C7 (2026-04-28)",
        "driver": "tools/test-mcp/e2e_phase_f.py --language cs",
        "intent": (
            "Vytvoř pro mě denního asistenta, který každé ráno v 8:00 shrne mé "
            "poznámky z předchozího dne uložené ve složce na mém disku. Žádné "
            "ruční schvalování — výstup zobraz v lište stavu. Bez paměti mezi "
            "dny."
        ),
        "steps": common_steps() + [
            ("Acceptance: localised field contains target-language diacritics (cz: čřšž etc.)", "ok"),
            ("Acceptance: UC ids contain NO target-language diacritics (English-stable)", "ok"),
            ("Acceptance: trigger_type strings contain NO target-language diacritics", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase F — Multi-language build (es / de / fr presets)",
        "summary": "Same gates as the Czech preset, target language varies. `--strict-language-percentage 0.5` raises the diacritic-density bar.",
        "introduced": "C7 (2026-04-28)",
        "driver": "tools/test-mcp/e2e_phase_f.py --language {es|de|fr}",
        "intent": (
            "[es] Crea un asistente matutino que cada mañana a las 8:00 me "
            "resuma las notas que añadí ayer en una carpeta de mi disco. Sin "
            "revisión humana — muestra el resultado en la barra de estado. Sin "
            "memoria entre días.\n"
            "\n"
            "[de] Erstelle mir einen morgendlichen Assistenten, der jeden "
            "Morgen um 8 Uhr meine Notizen vom Vortag aus einem Ordner auf "
            "meinem Laufwerk zusammenfasst. Keine manuelle Prüfung — zeige das "
            "Ergebnis in der Statusleiste. Kein Gedächtnis zwischen den Tagen.\n"
            "\n"
            "[fr] Crée-moi un assistant matinal qui chaque matin à 8h résume "
            "mes notes de la veille stockées dans un dossier de mon disque. "
            "Aucune revue manuelle — affiche le résultat dans la barre de "
            "statut. Sans mémoire entre les jours."
        ),
        "steps": [
            ("Run preset for `es` (Spanish)", "ok"),
            ("Run preset for `de` (German)", "ok"),
            ("Run preset for `fr` (French)", "ok"),
            ("All three localised in target language; identifiers stay English", "ok"),
        ],
    },
    {
        "title": "Phase G — Dry-run preview",
        "summary": "Tiny single-capability persona; calls `simulateBuildDraft` + `getSimulationArtefacts` WITHOUT promoting. Verifies wiring correctness.",
        "introduced": "C7 (2026-04-28)",
        "driver": "tools/test-mcp/e2e_phase_g.py",
        "intent": (
            "A simple notes summarizer. One capability: every morning at 8am "
            "local time, take any notes I added to a local-drive folder "
            "yesterday and produce a one-paragraph digest. Auto-publish (no "
            "review). Stateless."
        ),
        "steps": common_steps()[:6] + [  # no promote step — dry-run is pre-promote
            ("Resolve first capability id via getPersonaDetail", "ok"),
            ("simulateBuildDraft({useCaseId}) — backend snapshots design_context, dispatches simulate", "ok"),
            ("getSimulationArtefacts({executionId}) returns {executionId, reviews[], memories[]}", "ok"),
            ("Acceptance: response shape (executionId round-trips, both arrays are arrays)", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase H — Webhook trigger + smee auto-bind",
        "summary": "GitHub push → Slack relay scenario. Watches for `acceptsWebhookSource: true` clarifying-question, attaches per-run unique smee URL via typed-payload bridge call.",
        "introduced": "C7 (2026-04-28)",
        "driver": "tools/test-mcp/e2e_phase_h.py",
        "intent": (
            "React in real time when GitHub fires a `push` webhook on my main "
            "repo. Forward the payload to a Slack channel as a brief one-line "
            "summary so the team sees the commit immediately. No batching, no "
            "schedule — webhook trigger only."
        ),
        "steps": common_steps() + [
            ("During Q&A: detect `acceptsWebhookSource: true` question", "ok"),
            ("Submit smee URL via answerBuildQuestionWithWebhookSource (typed payload)", "ok"),
            ("Acceptance: IR has at least one webhook trigger", "ok"),
            ("Acceptance: webhook trigger config carries `smee_channel_url` matching submitted URL", "ok"),
            ("Acceptance: smee_relays row exists with target_persona_id == new persona, channel_url + event_filter match", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase I — Clockify monthly-invoice scenario",
        "summary": "2-capability persona (Generate Invoice + Prepare Email Draft) sharing a monthly schedule trigger. UC1 review_policy=never, UC2=always. 7 acceptance gates.",
        "introduced": "C7 (2026-04-28)",
        "driver": "tools/test-mcp/e2e_phase_i.py",
        "intent": (
            "Generate monthly invoices from my Clockify time entries. On the "
            "1st of each month at 9am, read last month's billable entries from "
            "Clockify, compose an itemized invoice that matches my saved "
            "template (I'll attach a sample), then prepare a draft email to my "
            "accountant with the invoice attached. Two capabilities: (1) "
            "Generate Invoice — produces the invoice document from Clockify "
            "data; (2) Prepare Email Draft — composes the message body and "
            "attaches the invoice. I want to review the email draft before it "
            "leaves my account."
        ),
        "steps": common_steps() + [
            ("G2: 2 use_cases produced", "ok"),
            ("G3: schedule trigger present (cron `0 9 1 * *`)", "ok"),
            ("G4: clockify connector visible (persona-level OR per-UC tool_hints)", "ok"),
            ("G5: email connector visible (gmail / resend / sendgrid / etc.)", "ok"),
            ("G6: at least one UC has human review (always or on_low_confidence)", "ok"),
            ("G7: design_context.useCases[i].review_policy carried through promote", "ok"),
            ("G9: last_design_result.persona.decision_principles non-empty (C7 hoist fix)", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase J — Documentation archiver",
        "summary": "2-UC scenario combining webhook+smee (Phase H plumbing) and reference attachment (Phase I plumbing, opportunistic). UC1 webhook-driven KB ingest, UC2 on-demand digest.",
        "introduced": "C8 (2026-04-28)",
        "driver": "tools/test-mcp/e2e_phase_j.py",
        "intent": (
            "Documentation archiver. Two capabilities. UC1: when GitHub fires "
            "a `push` webhook (forwarded via https://smee.io/phase-j-<random-hex>, "
            "event_filter `github.push`), parse the changed markdown files and "
            "store each as a KB fact. UC2: on demand only, output a short "
            "digest of recent KB facts. Auto-publish both — no human review.\n"
            "\n"
            "[Note: the `<random-hex>` segment is regenerated per run via "
            "uuid.uuid4().hex[:12] so re-runs don't collide on the smee_relays "
            "channel_url UNIQUE constraint. Pasting the URL verbatim hits rule "
            "24's SKIP path and bypasses the typed-payload question.]"
        ),
        "steps": common_steps() + [
            ("During Q&A: typed-payload listener for acceptsWebhookSource (with smee URL pasted in INTENT, rule 24 SKIP path can fire instead)", "ok"),
            ("Acceptance: at least 2 use_cases", "ok"),
            ("Acceptance: webhook trigger present", "ok"),
            ("Acceptance: webhook trigger config carries smee_channel_url matching INTENT", "ok"),
            ("Acceptance: smee_relays row exists with target_persona_id matching", "ok"),
            ("Acceptance: digest UC has non-webhook trigger (manual / no row)", "ok"),
            ("Latest run: smee_relays_created=1, all 5 gates green", "ok"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
    {
        "title": "Phase K — Video narration (build-shape only)",
        "summary": "Verifies build pipeline composes a narrated-video persona enumerating google_gemini (vision) + elevenlabs (TTS). Runtime blocked on TTS impl + ffmpeg connector.",
        "introduced": "C8 (2026-04-28)",
        "driver": "tools/test-mcp/e2e_phase_k.py",
        "intent": (
            "Make a narrated video on demand. I drop a folder of image frames "
            "into local-drive. For each frame, use Gemini Vision to write a "
            "one-sentence caption describing what's happening. Then use "
            "ElevenLabs to turn the captions into spoken audio. Final output: "
            "a video file with frames + spoken narration. Manual / on-demand "
            "trigger only — I'll invoke when frames are ready. Auto-publish, "
            "no human review."
        ),
        "steps": common_steps() + [
            ("Acceptance: at least one use_case", "ok"),
            ("Acceptance: design_context references google_gemini connector", "ok"),
            ("Acceptance: design_context references elevenlabs connector", "ok"),
            ("Acceptance: NO webhook trigger (scenario is on-demand)", "ok"),
            ("Latest run: 1 UC `Video Narrator`, manual trigger, both connectors landed", "ok"),
            ("Runtime execution (synthesise audio + compose video): not implementable today", "deferred"),
            ("Cleanup: deleteAgent succeeds", "ok"),
        ],
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

TITLE_FILL = PatternFill("solid", fgColor="1F2937")           # slate-800
TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")          # gray-200
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="111827")
SUMMARY_FONT = Font(name="Calibri", size=10, italic=True, color="4B5563")  # gray-600
BODY_FONT = Font(name="Calibri", size=11, color="111827")
INTENT_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="4338CA")  # indigo-700
INTENT_BODY_FONT = Font(name="Consolas", size=10, color="1F2937")             # mono — preserves prompt shape
INTENT_FILL = PatternFill("solid", fgColor="EEF2FF")                           # indigo-50

STATUS_FILLS = {
    "ok": PatternFill("solid", fgColor="D1FAE5"),                  # emerald-100
    "deferred": PatternFill("solid", fgColor="FEF3C7"),            # amber-100
    "deferred-manual": PatternFill("solid", fgColor="FED7AA"),     # orange-200
    "flaky": PatternFill("solid", fgColor="FEE2E2"),               # red-100
    "red": PatternFill("solid", fgColor="FCA5A5"),                 # red-300
}
STATUS_FONTS = {
    "ok": Font(name="Calibri", size=11, bold=True, color="065F46"),
    "deferred": Font(name="Calibri", size=11, bold=True, color="92400E"),
    "deferred-manual": Font(name="Calibri", size=11, bold=True, color="9A3412"),
    "flaky": Font(name="Calibri", size=11, bold=True, color="991B1B"),
    "red": Font(name="Calibri", size=11, bold=True, color="7F1D1D"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def write_workbook(dest: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Scenarios"

    # Column widths — first is the long flow-step description, second is the
    # short status chip, third is a wider notes column for ad-hoc manual notes.
    ws.column_dimensions["A"].width = 96
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60

    row = 1

    # Top-of-sheet legend
    ws.cell(row=row, column=1, value="Phase scenario suite — manual comparison checklist").font = Font(
        name="Calibri", size=14, bold=True, color="111827"
    )
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=(
            "Generated by tools/test-mcp/generate_test_scenarios_xlsx.py. Each "
            "scenario lists its canonical flow steps and the latest run status. "
            "Use the rightmost column for ad-hoc manual notes (false-greens, "
            "skipped steps, environment quirks). Status colours: green=ok, "
            "amber=deferred, orange=deferred-manual, red=flaky/failing."
        ),
    ).font = SUMMARY_FONT
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 60
    row += 2

    for scenario in SCENARIOS:
        # Title row (merged across A:C)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value=scenario["title"])
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
        cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1

        # Summary row (merged across A:C)
        summary = (
            f"{scenario['summary']}  ·  Introduced in {scenario['introduced']}  ·  "
            f"Driver: `{scenario['driver']}`"
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        sc = ws.cell(row=row, column=1, value=summary)
        sc.font = SUMMARY_FONT
        sc.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.row_dimensions[row].height = 32
        row += 1

        # Intent / prompt row (merged across A:C). Lets the user copy the
        # exact text the driver passes to startBuildFromIntent so they can
        # replicate the run by hand against the desktop app.
        intent_text = scenario.get("intent")
        if intent_text:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ic = ws.cell(row=row, column=1, value=f"Intent (prompt used):\n{intent_text}")
            ic.fill = INTENT_FILL
            ic.font = INTENT_BODY_FONT
            ic.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            ic.border = THIN_BORDER
            # Sized to fit ~the longest intent without truncation; xlsx
            # readers honour wrap_text, so taller intents grow naturally
            # when row height is set generously.
            line_count = max(intent_text.count("\n") + 1, 1)
            # Heuristic: 14 px per visual line. Cap so a long Phase B / D2
            # entry doesn't dwarf the rest of the sheet.
            ws.row_dimensions[row].height = min(14 + line_count * 14 + len(intent_text) // 12, 320)
            row += 1

        # Header row
        for col, label in enumerate(["Flow step", "Status", "Notes"], start=1):
            hc = ws.cell(row=row, column=col, value=label)
            hc.fill = HEADER_FILL
            hc.font = HEADER_FONT
            hc.alignment = Alignment(vertical="center", indent=1)
            hc.border = THIN_BORDER
        ws.row_dimensions[row].height = 18
        row += 1

        # Data rows
        for step, status in scenario["steps"]:
            sc = ws.cell(row=row, column=1, value=step)
            sc.font = BODY_FONT
            sc.alignment = Alignment(wrap_text=True, vertical="top")
            sc.border = THIN_BORDER

            tc = ws.cell(row=row, column=2, value=status)
            tc.fill = STATUS_FILLS.get(status, STATUS_FILLS["deferred"])
            tc.font = STATUS_FONTS.get(status, STATUS_FONTS["deferred"])
            tc.alignment = Alignment(horizontal="center", vertical="center")
            tc.border = THIN_BORDER

            nc = ws.cell(row=row, column=3, value="")  # blank for the user
            nc.border = THIN_BORDER
            ws.row_dimensions[row].height = 18
            row += 1

        # Blank separator
        row += 1

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    print(f"Wrote {dest} ({len(SCENARIOS)} scenarios)")


if __name__ == "__main__":
    write_workbook(DEST)
